from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Alignment

driver=webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()
driver.get("https://www.castrol.com/en_in/india/home/motorcycle-oil-and-fluids/bike-point/bikepoint-locations/motorcycle-workshop-delhi.html")
driver.implicitly_wait(20)
driver.find_element(By.XPATH,"//button[contains(@class,'nr-cookie-notification__cta-button nr-cookie-accept')]").click()
dealers=driver.find_element(By.XPATH,"//body/div[5]/div[3]/div[1]/div[1]/div[1]")#Delhi

mydealer=[]
mydealer.append(dealers.text)
finallist=zip(mydealer)
wb=Workbook()
wb['Sheet'].title='Delhi Dealers'
sh1=wb.active
for x in list(finallist):
    sh1.append(x)
sh1['A1'].alignment = Alignment(wrapText=True)
print("Delhi Done")

driver.find_element(By.XPATH,"//a[normalize-space()='MOTORCYCLE WORKSHOPS - BANGALORE']").click()
dealers1=driver.find_element(By.XPATH,"//body/div[5]/div[3]/div[1]/div[1]/div[1]")#Bangalore
mydealer1=[]
mydealer1.append(dealers1.text)
finallist1=zip(mydealer1)
ws=wb.create_sheet("Bangalore Dealers")
for y in list(finallist1):
    ws.append(y)
ws['A1'].alignment = Alignment(wrapText=True)
print("Bangalore Done")


driver.find_element(By.XPATH,"//a[normalize-space()='MOTORCYCLE WORKSHOPS - JAIPUR']").click()
dealers2=driver.find_element(By.XPATH,"//body/div[5]/div[3]/div[1]/div[1]/div[1]")#Jaipur
mydealer2=[]
mydealer2.append(dealers2.text)
finallist2=zip(mydealer2)
ws1=wb.create_sheet("Jaipur Dealers")
for y in list(finallist2):
    ws1.append(y)
ws1['A1'].alignment = Alignment(wrapText=True)
print("Jaipur Done")



driver.find_element(By.XPATH,"//a[normalize-space()='MOTORCYCLE WORKSHOPS - KOLKATA']").click()
dealers3=driver.find_element(By.XPATH,"//body/div[5]/div[3]/div[1]/div[1]/div[1]")#Kolkata
mydealer3=[]
mydealer3.append(dealers3.text)
finallist3=zip(mydealer3)
ws2=wb.create_sheet("Kolkata Dealers")
for y in list(finallist3):
    ws2.append(y)
ws2['A1'].alignment = Alignment(wrapText=True)
print("Kolkata Done")



driver.find_element(By.XPATH,"//a[normalize-space()='MOTORCYCLE WORKSHOP - MUMBAI']").click()
dealers4=driver.find_element(By.XPATH,"//body/div[5]/div[3]/div[1]/div[1]/div[1]")#Mumbai
mydealer4=[]
mydealer4.append(dealers4.text)
finallist4=zip(mydealer4)
ws3=wb.create_sheet("Mumbai Dealers")
for y in list(finallist4):
    ws3.append(y)
print("Mumbai Done")
ws3['A1'].alignment = Alignment(wrapText=True)
wb.save("FinalCastrol.xlsx")